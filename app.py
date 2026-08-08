from datetime import date, datetime, timedelta
import json
import re
import sqlite3
import urllib.error
import urllib.parse
import urllib.request
import os

from flask import Flask, g, redirect, render_template, request, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash

# ── 資料庫設定 ──────────────────────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL")  # Render PostgreSQL
DATABASE_SQLITE = "entries.db"                 # 本機 SQLite fallback

# 這段邏輯會自動處理 Supabase 的連線字串格式
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-in-production-please")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "請先登入才能使用此功能。"

VALID_TICKER_RE = re.compile(r"^[A-Za-z0-9\.\-]{1,10}$")


# ── 使用者模型 ───────────────────────────────────────────────
class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username


@login_manager.user_loader
def load_user(user_id):
    rows = db_fetchall(
        "SELECT id, username FROM users WHERE id = %s" if USE_POSTGRES else "SELECT id, username FROM users WHERE id = ?",
        (int(user_id),),
    )
    if rows:
        return User(rows[0]["id"], rows[0]["username"])
    return None


# ── 資料庫連線 ───────────────────────────────────────────────
def get_db():
    db = getattr(g, "db", None)
    if db is None:
        if USE_POSTGRES:
            db = psycopg2.connect(DATABASE_URL)
            g.db = db
        else:
            db = sqlite3.connect(DATABASE_SQLITE)
            db.row_factory = sqlite3.Row
            g.db = db
    return db


@app.teardown_appcontext
def close_db(exc):
    db = getattr(g, "db", None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()
    if USE_POSTGRES:
        with db.cursor() as cur:
            # 建立 users 資料表
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )
            # 建立 entries 資料表（含 user_id）
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS entries (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES users(id),
                    trade_date TEXT NOT NULL,
                    ticker TEXT NOT NULL,
                    shares REAL NOT NULL,
                    close_price REAL NOT NULL,
                    usd_twd REAL NOT NULL,
                    value_usd REAL,
                    value_twd REAL NOT NULL,
                    source TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )
        # 若 entries 已存在但缺少 user_id 欄位，補上
        with db.cursor() as cur:
            cur.execute("""
                ALTER TABLE entries ADD COLUMN IF NOT EXISTS user_id INTEGER REFERENCES users(id)
            """)
        db.commit()
    else:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                trade_date TEXT NOT NULL,
                ticker TEXT NOT NULL,
                shares REAL NOT NULL,
                close_price REAL NOT NULL,
                usd_twd REAL NOT NULL,
                value_usd REAL,
                value_twd REAL NOT NULL,
                source TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        cur = db.execute("PRAGMA table_info(entries)")
        cols = [r[1] for r in cur.fetchall()]
        if "value_usd" not in cols:
            try:
                db.execute("ALTER TABLE entries ADD COLUMN value_usd REAL")
            except Exception:
                pass
        if "user_id" not in cols:
            try:
                db.execute("ALTER TABLE entries ADD COLUMN user_id INTEGER")
            except Exception:
                pass
        db.commit()


# ── 共用查詢輔助函式 ─────────────────────────────────────────
def db_fetchall(query, params=()):
    db = get_db()
    if USE_POSTGRES:
        with db.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(query, params)
            return cur.fetchall()
    else:
        return db.execute(query, params).fetchall()


def db_execute(query, params=()):
    db = get_db()
    if USE_POSTGRES:
        with db.cursor() as cur:
            cur.execute(query, params)
        db.commit()
    else:
        db.execute(query, params)
        db.commit()


def db_placeholder(n):
    """回傳對應數量的佔位符，PostgreSQL 用 %s，SQLite 用 ?"""
    ph = "%s" if USE_POSTGRES else "?"
    return ", ".join([ph] * n)


# ── 原有功能函式（不變）──────────────────────────────────────


def parse_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("日期格式須為 yyyy-mm-dd")


def http_get_json(url, headers=None):
    req = urllib.request.Request(url, headers=headers or {"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=15) as response:
        raw = response.read().decode("utf-8", errors="ignore")
    return json.loads(raw)


def fetch_yahoo_close_price(symbol, target_date):
    start_dt = datetime.combine(target_date - timedelta(days=10), datetime.min.time())
    end_dt = datetime.combine(target_date + timedelta(days=1), datetime.min.time())
    period1 = int(start_dt.timestamp())
    period2 = int(end_dt.timestamp())
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}"
        f"?period1={period1}&period2={period2}&interval=1d&includePrePost=false"
    )
    data = http_get_json(url)
    result = data.get("chart", {}).get("result")
    if not result:
        raise ValueError(f"無法取得 {symbol} 的資料")
    quote = result[0].get("indicators", {}).get("quote", [])
    timestamps = result[0].get("timestamp", [])
    if not quote or not timestamps:
        raise ValueError(f"{symbol} 無歷史交易資料")
    closes = quote[0].get("close", [])
    best_price = None
    best_timestamp = None
    for ts, close in zip(timestamps, closes):
        if close is None:
            continue
        current = date.fromtimestamp(ts)
        if current <= target_date:
            if best_timestamp is None or current > date.fromtimestamp(best_timestamp):
                best_timestamp = ts
                best_price = close
    if best_price is None:
        raise ValueError(f"{symbol} 在指定日或之前沒有可用收盤價")
    return best_price, date.fromtimestamp(best_timestamp)


def fetch_usd_twd_rate_for_today():
    url = "https://rate.bot.com.tw/xrt?Lang=zh-TW"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=15) as response:
        html = response.read().decode("utf-8", errors="ignore")
    usd_rows = re.findall(r"<tr.*?>.*?</tr>", html, re.S)
    for row in usd_rows:
        if "美金" in row and "(USD)" in row:
            numbers = re.findall(r"<td[^>]*>\s*([0-9.,]+)\s*</td>", row)
            if numbers:
                buy_rate = numbers[0].replace(",", "")
                return float(buy_rate)
    raise ValueError("無法解析台灣銀行匯率資料")


def get_previous_trading_date(target_date):
    """取得指定日期的前一個交易日（跳過週末，假日由 Yahoo 自動處理）"""
    prev = target_date - timedelta(days=1)
    # 跳過週六(5)和週日(6)
    while prev.weekday() >= 5:
        prev -= timedelta(days=1)
    return prev


def get_prev_trading_day(d):
    """
    取得指定日期的前一個交易日：
    - 平日(週一~週五) → 往前1天
    - 週六 → 往前2天
    - 週日 → 往前3天
    結果若仍為週末則繼續往前（保險用）
    """
    wd = d.weekday()
    if wd == 5:    # 週六
        d -= timedelta(days=2)
    elif wd == 6:  # 週日
        d -= timedelta(days=3)
    else:          # 週一~週五
        d -= timedelta(days=1)
    # 保險：若結果仍為週末繼續往前
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def fetch_usd_twd_rate(symbol, target_date):
    if target_date == date.today():
        try:
            return fetch_usd_twd_rate_for_today(), "台灣銀行本行買入"
        except Exception:
            pass
    price, _actual_date = fetch_yahoo_close_price(symbol, target_date)
    return price, "Yahoo 匯率"


def get_last_trading_day_before(d):
    """取得指定日期前一個交易日（若為週末則往前跳到週五再減一天到週四）"""
    # 先跳到當天或之前的最近平日
    while d.weekday() >= 5:  # 5=週六, 6=週日
        d -= timedelta(days=1)
    # 再往前一天，確保是「之前」的交易日而非當天
    d -= timedelta(days=1)
    # 若往前一天又落在週末，繼續往前
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def get_espp_reference_dates(purchase_date):
    """
    根據 ESPP 認購日（1/31 或 7/31）回傳參考日期
    - 1/31：參考A=當年 1/31 前一交易日，參考B=前一年 8/1 前一交易日
    - 7/31：參考A=當年 7/31 前一交易日，參考B=當年 2/1 前一交易日
    回傳: (price_date_a, price_date_b, rate_date)
    """
    year = purchase_date.year
    month = purchase_date.month

    if month == 1:  # 1/31 認購
        price_date_a = get_prev_trading_day(date(year, 1, 31))
        price_date_b = get_prev_trading_day(date(year - 1, 8, 1))  # 前一年 8/1
        rate_date = get_prev_trading_day(date(year, 1, 31))
    elif month == 7:  # 7/31 認購
        price_date_a = get_prev_trading_day(date(year, 7, 31))
        price_date_b = get_prev_trading_day(date(year, 2, 1))      # 當年 2/1
        rate_date = get_prev_trading_day(date(year, 7, 31))
    else:
        raise ValueError("ESPP 認購日只能是 1/31 或 7/31")

    return price_date_a, price_date_b, rate_date


# ── 使用者路由 ────────────────────────────────────────────────
@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()
        if not username or not password:
            error = "請填寫帳號與密碼。"
        elif len(password) < 6:
            error = "密碼至少需要 6 個字元。"
        elif password != confirm:
            error = "兩次密碼輸入不一致。"
        else:
            existing = db_fetchall(
                "SELECT id FROM users WHERE username = %s" if USE_POSTGRES else "SELECT id FROM users WHERE username = ?",
                (username,),
            )
            if existing:
                error = "此帳號已被使用，請換一個。"
            else:
                ph = db_placeholder(3)
                db_execute(
                    f"INSERT INTO users (username, password_hash, created_at) VALUES ({ph})",
                    (username, generate_password_hash(password), datetime.now().isoformat(timespec="seconds")),
                )
                flash("註冊成功！請登入。", "success")
                return redirect(url_for("login"))
    return render_template("register.html", error=error)


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        rows = db_fetchall(
            "SELECT id, username, password_hash FROM users WHERE username = %s" if USE_POSTGRES else "SELECT id, username, password_hash FROM users WHERE username = ?",
            (username,),
        )
        if not rows or not check_password_hash(rows[0]["password_hash"], password):
            error = "帳號或密碼錯誤。"
        else:
            user = User(rows[0]["id"], rows[0]["username"])
            login_user(user)
            return redirect(url_for("index"))
    return render_template("login.html", error=error)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("已登出。", "info")
    return redirect(url_for("login"))


# ── 路由 ─────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    rsu_error = None
    rsu_message = None
    espp_error = None
    espp_message = None
    espp_result = None
    active_tab = request.form.get("active_tab", "rsu")  # 記住目前 tab

    if request.method == "POST":
        form_type = request.form.get("form_type", "rsu")
        active_tab = form_type

        if form_type == "rsu":
            trade_date = request.form.get("trade_date", "").strip()
            ticker = request.form.get("ticker", "").strip().upper()
            shares = request.form.get("shares", "").strip()
            if not trade_date or not ticker or not shares:
                rsu_error = "請填寫完整的日期、美股代號與股數。"
            elif not VALID_TICKER_RE.match(ticker):
                rsu_error = "股票代號格式不正確。"
            else:
                try:
                    trade_date_obj = parse_date(trade_date)
                    if trade_date_obj > date.today():
                        raise ValueError("日期錯誤：不能選擇未來日期")
                    shares_value = float(shares)
                    if shares_value <= 0:
                        raise ValueError("股數必須大於 0")
                    price_date = get_previous_trading_date(trade_date_obj)
                    close_price, stock_date = fetch_yahoo_close_price(ticker, price_date)
                    usd_twd_rate, rate_source = fetch_usd_twd_rate("USDTWD=X", price_date)
                    value_twd = round(close_price * shares_value * usd_twd_rate, 2)
                    value_usd = round(close_price * shares_value, 4)
                    ph = db_placeholder(10)
                    db_execute(
                        f"INSERT INTO entries (user_id, trade_date, ticker, shares, close_price, usd_twd, value_usd, value_twd, source, created_at) VALUES ({ph})",
                        (
                            current_user.id,
                            trade_date_obj.isoformat(),
                            ticker,
                            shares_value,
                            round(close_price, 4),
                            round(usd_twd_rate, 4),
                            value_usd,
                            value_twd,
                            f"RSU 股票 {ticker}({stock_date}) / 匯率 {rate_source}",
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    rsu_message = f"已新增 {ticker} 的計算結果，總價值 {value_twd:,} TWD。"
                except Exception as exc:
                    rsu_error = str(exc)

        elif form_type == "espp":
            purchase_date_str = request.form.get("purchase_date", "").strip()
            ticker = request.form.get("ticker", "").strip().upper()
            shares = request.form.get("shares", "").strip()
            if not purchase_date_str or not ticker or not shares:
                espp_error = "請填寫完整的認購日期、美股代號與股數。"
            elif not VALID_TICKER_RE.match(ticker):
                espp_error = "股票代號格式不正確。"
            else:
                try:
                    purchase_date = parse_date(purchase_date_str)
                    if purchase_date.day != 31 or purchase_date.month not in (1, 7):
                        raise ValueError("ESPP 認購日必須是 1/31 或 7/31")
                    shares_value = float(shares)
                    if shares_value <= 0:
                        raise ValueError("股數必須大於 0")
                    price_date_a, price_date_b, rate_date = get_espp_reference_dates(purchase_date)
                    price_a, actual_date_a = fetch_yahoo_close_price(ticker, price_date_a)
                    price_b, actual_date_b = fetch_yahoo_close_price(ticker, price_date_b)
                    cost_price = round(min(price_a, price_b) * 0.85, 4)
                    income_price = round(price_a, 4)  # 所得價固定為參考價格A
                    usd_twd_rate, rate_source = fetch_usd_twd_rate("USDTWD=X", rate_date)
                    gain_per_share = round(income_price - cost_price, 4)
                    value_usd = round(gain_per_share * shares_value, 4)
                    value_twd = round(value_usd * usd_twd_rate, 2)
                    espp_result = {
                        "ticker": ticker,
                        "purchase_date": purchase_date_str,
                        "price_date_a": actual_date_a.isoformat(),
                        "price_date_b": actual_date_b.isoformat(),
                        "price_a": price_a,
                        "price_b": price_b,
                        "cost_price": cost_price,
                        "income_price": income_price,
                        "gain_per_share": gain_per_share,
                        "shares": shares_value,
                        "usd_twd_rate": usd_twd_rate,
                        "rate_source": rate_source,
                        "value_usd": value_usd,
                        "value_twd": value_twd,
                        "value_usd_formatted": f"{value_usd:,.2f}",
                        "value_twd_formatted": f"{value_twd:,.2f}",
                    }
                    ph = db_placeholder(10)
                    db_execute(
                        f"INSERT INTO entries (user_id, trade_date, ticker, shares, close_price, usd_twd, value_usd, value_twd, source, created_at) VALUES ({ph})",
                        (
                            current_user.id,
                            purchase_date_str,
                            ticker,
                            shares_value,
                            income_price,
                            round(usd_twd_rate, 4),
                            value_usd,
                            value_twd,
                            f"ESPP {ticker} 所得價{income_price}({actual_date_a}) / 成本價{cost_price}({actual_date_b}) / 匯率 {rate_source}",
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    espp_message = f"ESPP 計算完成，所得價值 {value_twd:,.2f} TWD。"
                except Exception as exc:
                    espp_error = str(exc)

    ph = "%s" if USE_POSTGRES else "?"
    rows_raw = db_fetchall(f"SELECT * FROM entries WHERE user_id = {ph} ORDER BY id DESC", (current_user.id,))
    processed_rows = []
    for row in rows_raw:
        value_twd = float(row["value_twd"]) if row["value_twd"] is not None else 0.0
        value_usd = float(row["value_usd"]) if row["value_usd"] is not None else (value_twd / float(row["usd_twd"]) if row["usd_twd"] else 0.0)
        processed_rows.append(
            {
                "id": row["id"],
                "trade_date": row["trade_date"],
                "ticker": row["ticker"],
                "shares": row["shares"],
                "close_price": row["close_price"],
                "usd_twd": row["usd_twd"],
                "value_twd": value_twd,
                "value_usd": value_usd,
                "source": row["source"],
                "value_formatted": f"{value_twd:,.2f}",
                "value_usd_formatted": f"{value_usd:,.2f}",
            }
        )
    total = sum(r["value_twd"] for r in processed_rows) if processed_rows else 0.0
    total_formatted = f"{total:,.2f}"
    return render_template(
        "index.html",
        rows=processed_rows,
        total=round(total, 2),
        total_formatted=total_formatted,
        active_tab=active_tab,
        rsu_error=rsu_error,
        rsu_message=rsu_message,
        espp_error=espp_error,
        espp_message=espp_message,
        espp_result=espp_result,
    )


@app.route("/espp", methods=["GET", "POST"])
def espp():
    error = None
    message = None
    result = None

    if request.method == "POST":
        purchase_date_str = request.form.get("purchase_date", "").strip()
        ticker = request.form.get("ticker", "").strip().upper()
        shares = request.form.get("shares", "").strip()

        if not purchase_date_str or not ticker or not shares:
            error = "請填寫完整的認購日期、美股代號與股數。"
        elif not VALID_TICKER_RE.match(ticker):
            error = "股票代號格式不正確。"
        else:
            try:
                purchase_date = parse_date(purchase_date_str)
                if purchase_date.day not in (31,) or purchase_date.month not in (1, 7):
                    raise ValueError("ESPP 認購日必須是 1/31 或 7/31")
                shares_value = float(shares)
                if shares_value <= 0:
                    raise ValueError("股數必須大於 0")

                price_date_a, price_date_b, rate_date = get_espp_reference_dates(purchase_date)

                # 抓兩個參考日的股價
                price_a, actual_date_a = fetch_yahoo_close_price(ticker, price_date_a)
                price_b, actual_date_b = fetch_yahoo_close_price(ticker, price_date_b)

                # 計算成本價與所得價
                cost_price = round(min(price_a, price_b) * 0.85, 4)   # 較低價 × 85%
                income_price = round(price_a, 4)                        # 所得價固定為參考價格A

                # 抓匯率（認購日當日）
                usd_twd_rate, rate_source = fetch_usd_twd_rate("USDTWD=X", rate_date)

                # 計算價值
                gain_per_share = round(income_price - cost_price, 4)
                value_usd = round(gain_per_share * shares_value, 4)
                value_twd = round(value_usd * usd_twd_rate, 2)

                result = {
                    "ticker": ticker,
                    "purchase_date": purchase_date_str,
                    "price_date_a": actual_date_a.isoformat(),
                    "price_date_b": actual_date_b.isoformat(),
                    "price_a": price_a,
                    "price_b": price_b,
                    "cost_price": cost_price,
                    "income_price": income_price,
                    "gain_per_share": gain_per_share,
                    "shares": shares_value,
                    "usd_twd_rate": usd_twd_rate,
                    "rate_source": rate_source,
                    "value_usd": value_usd,
                    "value_twd": value_twd,
                    "value_usd_formatted": f"{value_usd:,.2f}",
                    "value_twd_formatted": f"{value_twd:,.2f}",
                }

                # 儲存到資料庫
                ph = db_placeholder(9)
                db_execute(
                    f"INSERT INTO entries (trade_date, ticker, shares, close_price, usd_twd, value_usd, value_twd, source, created_at) VALUES ({ph})",
                    (
                        purchase_date_str,
                        ticker,
                        shares_value,
                        income_price,
                        round(usd_twd_rate, 4),
                        value_usd,
                        value_twd,
                        f"ESPP {ticker} 所得價{income_price}({actual_date_a}) / 成本價{cost_price}({actual_date_b}) / 匯率 {rate_source}",
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )
                message = f"ESPP 計算完成，所得價值 {value_twd:,.2f} TWD。"

            except Exception as exc:
                error = str(exc)

    return render_template("espp.html", error=error, message=message, result=result)


@app.route("/delete/<int:entry_id>", methods=["POST"])
@login_required
def delete_entry(entry_id):
    ph = "%s" if USE_POSTGRES else "?"
    # 確保只能刪除自己的紀錄，防止刪除別人的資料
    db_execute(f"DELETE FROM entries WHERE id = {ph} AND user_id = {ph}", (entry_id, current_user.id))
    return redirect(url_for("index"))


@app.route("/clear", methods=["POST"])
@login_required
def clear_entries():
    ph = "%s" if USE_POSTGRES else "?"
    db_execute(f"DELETE FROM entries WHERE user_id = {ph}", (current_user.id,))
    return redirect(url_for("index"))


# ── 匯出功能 ─────────────────────────────────────────────────
import csv
import io
from flask import make_response

@app.route("/export/csv")
@login_required
def export_csv():
    ph = "%s" if USE_POSTGRES else "?"
    rows = db_fetchall(f"SELECT * FROM entries WHERE user_id = {ph} ORDER BY id DESC", (current_user.id,))
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["日期", "代號", "股數", "股價/所得價(USD)", "匯率(USD/TWD)", "價值(TWD)", "價值(USD)", "備註", "建立時間"])
    for r in rows:
        writer.writerow([
            r["trade_date"], r["ticker"], r["shares"],
            r["close_price"], r["usd_twd"],
            r["value_twd"], r["value_usd"],
            r["source"], r["created_at"]
        ])
    # 加上 UTF-8 BOM，確保 Excel 開啟中文不亂碼
    content = '\ufeff' + output.getvalue()
    response = make_response(content.encode('utf-8'))
    response.headers["Content-Disposition"] = "attachment; filename=rsu_espp_records.csv"
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    return response


@app.route("/export/excel")
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "請先安裝 openpyxl：pip install openpyxl", 500

    ph = "%s" if USE_POSTGRES else "?"
    rows = db_fetchall(f"SELECT * FROM entries WHERE user_id = {ph} ORDER BY id DESC", (current_user.id,))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RSU ESPP 紀錄"

    headers = ["日期", "代號", "股數", "股價/所得價(USD)", "匯率(USD/TWD)", "價值(TWD)", "價值(USD)", "備註", "建立時間"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["trade_date"])
        ws.cell(row=row_idx, column=2, value=r["ticker"])
        ws.cell(row=row_idx, column=3, value=float(r["shares"]))
        ws.cell(row=row_idx, column=4, value=float(r["close_price"]))
        ws.cell(row=row_idx, column=5, value=float(r["usd_twd"]))
        ws.cell(row=row_idx, column=6, value=float(r["value_twd"]))
        ws.cell(row=row_idx, column=7, value=float(r["value_usd"]) if r["value_usd"] else 0)
        ws.cell(row=row_idx, column=8, value=r["source"])
        ws.cell(row=row_idx, column=9, value=r["created_at"])

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=rsu_espp_records.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


# ── 應用程式啟動時初始化資料庫 ──────────────────────────────
with app.app_context():
    try:
        init_db()
        print("[startup] 資料庫初始化成功")
    except Exception as e:
        print(f"[startup] 資料庫初始化失敗：{e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)